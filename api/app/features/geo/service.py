"""The geography master: reading it, and loading it from a spreadsheet.

The importer is additive. It creates and updates what the file names and never
deletes what the file omits — a partial upload must not silently unmap half of
India. Everything it decides that a human might disagree with is reported:
re-parented rows and every rejected row, with its reason.

**The spreadsheet is the only source.** There are no hard-coded corrections. An
earlier version carried researched overrides for a dozen pincodes, and they had
to go: an override outranks the file, so fixing the file stopped fixing the
master, and nobody could tell from the sheet what the master would end up
holding. Corrections are made in the sheet and re-uploaded —
`RequirementDocs/apply-pincode-corrections.py` records the ones already applied.

Parsing is streamed (`read_only=True`), because the real file is 165,627 rows —
one per post office, roughly 8.5 rows per pincode.
"""

import csv
import io
import re
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field

from fastapi import HTTPException, status
from sqlalchemy import bindparam, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.schemas import ListParams
from app.features.geo.schemas import (
    DistrictOut,
    ImportCounts,
    ImportOverride,
    ImportReject,
    ImportReport,
    PincodeOut,
    RegionOut,
    StateOut,
)
from app.models.territory import District, Pincode, PincodeDistrict, Region, State

#: The file is 165,627 rows; this is headroom, not a squeeze.
MAX_ROWS = 200_000
#: Nothing goes to blob storage here — the 8 MB image ceiling does not apply.
MAX_UPLOAD_BYTES = 16 * 1024 * 1024
#: An Indian pincode never begins with 0.
PINCODE_RE = re.compile(r"^[1-9][0-9]{5}$")
#: How many rejects travel back in the response; the count is always exact.
MAX_REJECTS_RETURNED = 200

#: Header spellings we accept, lowercased with all non-letters stripped — so
#: "Pin Code", "PINCODE" and "pin_code" all land on the same column.
_HEADERS = {
    "region": "region",
    "state": "state",
    "district": "district",
    "pincode": "pincode",
    "pin": "pincode",
    "pincodes": "pincode",
}

#: Words that stay lowercase inside a title-cased name, so
#: "THE DADRA AND NAGAR HAVELI AND DAMAN AND DIU" reads like a place.
_MINOR_WORDS = {"and", "of", "the", "at", "in", "on"}

#: Rows the source could not resolve. They are dropped, never imported as a
#: region literally called "#N/A".
_SOURCE_ERRORS = {"#n/a", "na", "n/a", "#value!", "#ref!", "null", "-", ""}


def build_template() -> io.BytesIO:
    """A one-sheet .xlsx with the four headers and one example row."""
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Pincodes"
    sheet.append(["Region", "State", "District", "Pin Code"])
    sheet.append(["South", "TELANGANA", "HYDERABAD", 500001])
    for column, width in zip("ABCD", (14, 30, 26, 12)):
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def _norm_header(raw: object) -> str:
    return re.sub(r"[^a-z]", "", str(raw or "").lower())


def title_case(raw: str) -> str:
    """`TELANGANA` -> `Telangana`, keeping joining words lowercase."""
    words = [w for w in re.split(r"\s+", raw.strip()) if w]
    out: list[str] = []
    for i, w in enumerate(words):
        low = w.lower()
        if i > 0 and low in _MINOR_WORDS:
            out.append(low)
        elif "-" in w:
            out.append("-".join(p.capitalize() for p in low.split("-")))
        else:
            out.append(low.capitalize())
    return " ".join(out)


# ── reading ────────────────────────────────────────────────────────────────


async def list_regions(
    session: AsyncSession, *, region_ids: list[uuid.UUID] | None = None
) -> list[RegionOut]:
    """Every region, active or not, with what sits under it.

    Includes regions with ZERO states on purpose: a region nobody can usefully
    be assigned to is exactly what the Geography screen has to surface.

    `region_ids` narrows to a caller's own territory — resolved at the router,
    never here, because geography is a global master and this stays a query over
    a table nobody owns. `None` means no narrowing; an EMPTY list would mean "no
    regions" and the router returns early rather than sending one.

    All three counts are computed here rather than summed in the browser — the
    console would otherwise add up 36 state rows to draw five tiles, and would
    get the pincode figure wrong the moment a state is filtered out.
    """
    states = (
        select(State.region_id, func.count().label("n"))
        .group_by(State.region_id)
        .subquery()
    )
    districts = (
        select(State.region_id, func.count(District.id).label("n"))
        .join(District, District.state_id == State.id)
        .group_by(State.region_id)
        .subquery()
    )
    pincodes = (
        select(State.region_id, func.count(Pincode.code).label("n"))
        .join(Pincode, Pincode.state_id == State.id)
        .group_by(State.region_id)
        .subquery()
    )
    stmt = (
        select(
            Region,
            func.coalesce(states.c.n, 0),
            func.coalesce(districts.c.n, 0),
            func.coalesce(pincodes.c.n, 0),
        )
        .outerjoin(states, states.c.region_id == Region.id)
        .outerjoin(districts, districts.c.region_id == Region.id)
        .outerjoin(pincodes, pincodes.c.region_id == Region.id)
        .order_by(Region.sort_order)
    )
    if region_ids is not None:
        stmt = stmt.where(Region.id.in_(region_ids))
    rows = await session.execute(stmt)
    return [
        RegionOut(
            id=region.id,
            code=region.code,
            name=region.name,
            isActive=region.is_active,
            stateCount=int(scount),
            districtCount=int(dcount),
            pincodeCount=int(pcount),
        )
        for region, scount, dcount, pcount in rows
    ]


async def list_districts(
    session: AsyncSession,
    *,
    state_id: uuid.UUID | None = None,
    region_id: uuid.UUID | None = None,
    state_ids: list[uuid.UUID] | None = None,
    region_ids: list[uuid.UUID] | None = None,
) -> list[DistrictOut]:
    """Districts with their pincode counts. Unpaged — 754 in all, 75 at most
    in one state (Uttar Pradesh).

    Counted through `pincode_districts`, because a pincode has no district
    column: 1,209 of them span two to four districts and are counted in each.
    So these numbers deliberately do NOT sum to the state's pincode count, and
    nothing may present them as if they did.
    """
    counts = (
        select(PincodeDistrict.district_id, func.count().label("n"))
        .group_by(PincodeDistrict.district_id)
        .subquery()
    )
    stmt = (
        select(District, State, Region, func.coalesce(counts.c.n, 0))
        .join(State, State.id == District.state_id)
        .join(Region, Region.id == State.region_id)
        .outerjoin(counts, counts.c.district_id == District.id)
    )
    if state_id is not None:
        stmt = stmt.where(District.state_id == state_id)
    if region_id is not None:
        stmt = stmt.where(State.region_id == region_id)
    # The plural forms are a territory restriction, applied ON TOP of any
    # single-id filter above rather than instead of it — so asking for one
    # state you do not cover returns nothing, which is the honest answer.
    if state_ids is not None:
        stmt = stmt.where(District.state_id.in_(state_ids))
    if region_ids is not None:
        stmt = stmt.where(State.region_id.in_(region_ids))

    rows = await session.execute(
        stmt.order_by(Region.sort_order, State.name, District.name)
    )
    return [
        DistrictOut(
            id=district.id,
            name=district.name,
            stateId=state.id,
            stateName=state.name,
            regionId=region.id,
            regionName=region.name,
            pincodeCount=int(count),
        )
        for district, state, region, count in rows
    ]


async def list_states(
    session: AsyncSession,
    *,
    region_ids: list[uuid.UUID] | None = None,
    state_ids: list[uuid.UUID] | None = None,
) -> list[StateOut]:
    """Every state with its region and counts. 36 rows — deliberately unpaged.

    Narrowed the same way `list_regions` is, and by whichever side the caller
    actually holds: an area manager holds STATES, a regional head holds REGIONS
    and no states at all. `state_ids` wins when both are given, because it is
    the more specific of the two.
    """
    districts = (
        select(District.state_id, func.count().label("n"))
        .group_by(District.state_id)
        .subquery()
    )
    pincodes = (
        select(Pincode.state_id, func.count().label("n"))
        .group_by(Pincode.state_id)
        .subquery()
    )
    stmt = (
        select(
            State,
            Region,
            func.coalesce(districts.c.n, 0),
            func.coalesce(pincodes.c.n, 0),
        )
        .join(Region, Region.id == State.region_id)
        .outerjoin(districts, districts.c.state_id == State.id)
        .outerjoin(pincodes, pincodes.c.state_id == State.id)
        .order_by(Region.sort_order, State.name)
    )
    if state_ids is not None:
        stmt = stmt.where(State.id.in_(state_ids))
    elif region_ids is not None:
        stmt = stmt.where(State.region_id.in_(region_ids))
    rows = await session.execute(stmt)
    return [
        StateOut(
            id=state.id,
            name=state.name,
            regionId=region.id,
            regionName=region.name,
            isActive=state.is_active,
            districtCount=int(dcount),
            pincodeCount=int(pcount),
        )
        for state, region, dcount, pcount in rows
    ]


async def list_pincodes(
    session: AsyncSession,
    params: ListParams,
    *,
    state_id: uuid.UUID | None = None,
    region_id: uuid.UUID | None = None,
    district_id: uuid.UUID | None = None,
    no_district: bool = False,
) -> tuple[list[PincodeOut], int]:
    """Paginated pincodes. This is what a coverage picker searches."""
    stmt = (
        select(Pincode, State, Region)
        .join(State, State.id == Pincode.state_id)
        .join(Region, Region.id == State.region_id)
    )
    if state_id is not None:
        stmt = stmt.where(Pincode.state_id == state_id)
    if region_id is not None:
        stmt = stmt.where(State.region_id == region_id)
    if district_id is not None:
        # EXISTS rather than a join: the composite primary key means a join on
        # one district could not duplicate a row today, but a semi-join cannot
        # start duplicating them later either, and `total` below counts this
        # same statement.
        stmt = stmt.where(
            select(PincodeDistrict.pincode_code)
            .where(
                PincodeDistrict.pincode_code == Pincode.code,
                PincodeDistrict.district_id == district_id,
            )
            .exists()
        )
    if no_district:
        # Four real pincodes have no district link at all. Without a way to ask
        # for them they are unreachable from a district drill-down — visible in
        # the state's total and in none of its districts, which reads as a bug.
        stmt = stmt.where(
            ~select(PincodeDistrict.pincode_code)
            .where(PincodeDistrict.pincode_code == Pincode.code)
            .exists()
        )
    if params.search:
        term = f"%{params.search.lower()}%"
        # District too, not just code and state: the picker labels each row
        # "500001 - Hyderabad", so a district is the obvious thing to type and
        # finding nothing would read as "we don't serve Hyderabad".
        in_district = (
            select(PincodeDistrict.pincode_code)
            .join(District, District.id == PincodeDistrict.district_id)
            .where(
                PincodeDistrict.pincode_code == Pincode.code,
                func.lower(District.name).like(term),
            )
            .exists()
        )
        stmt = stmt.where(
            or_(
                Pincode.code.like(f"{params.search.strip()}%"),
                func.lower(State.name).like(term),
                in_district,
            )
        )
    stmt = stmt.order_by(Pincode.code)

    total = await session.scalar(
        select(func.count()).select_from(stmt.order_by(None).subquery())
    )
    rows = (
        await session.execute(
            stmt.limit(params.limit).offset((params.page - 1) * params.limit)
        )
    ).all()

    codes = [p.code for p, _s, _r in rows]
    names: dict[str, list[str]] = defaultdict(list)
    if codes:
        joined = await session.execute(
            select(PincodeDistrict.pincode_code, District.name)
            .join(District, District.id == PincodeDistrict.district_id)
            .where(PincodeDistrict.pincode_code.in_(codes))
            .order_by(District.name)
        )
        for code, name in joined:
            names[code].append(name)

    return [
        PincodeOut(
            code=p.code,
            stateId=s.id,
            stateName=s.name,
            regionId=r.id,
            regionName=r.name,
            districts=names.get(p.code, []),
        )
        for p, s, r in rows
    ], int(total or 0)


# ── parsing ────────────────────────────────────────────────────────────────


@dataclass
class _Parsed:
    """What the sheet says, after cleaning but before anything is written."""

    rows_read: int = 0
    rows_skipped: int = 0
    #: state name (as written) -> region name
    state_region: dict[str, str] = field(default_factory=dict)
    #: pincode -> Counter of state names, so a conflict can be judged by weight
    pin_states: dict[str, Counter] = field(default_factory=lambda: defaultdict(Counter))
    #: pincode -> {district name}
    pin_districts: dict[str, set[str]] = field(
        default_factory=lambda: defaultdict(set)
    )
    #: state name -> {district name}
    state_districts: dict[str, set[str]] = field(
        default_factory=lambda: defaultdict(set)
    )
    #: Well-formed pincodes seen ONLY on rows the source could not resolve. Most
    #: also appear on a good row and are placed normally; the leftovers are
    #: reported by name rather than vanishing into the skipped count.
    unresolved_pins: set[str] = field(default_factory=set)
    rejects: list[ImportReject] = field(default_factory=list)


def _iter_rows(data: bytes, filename: str):
    """Yield raw tuples from .xlsx or .csv, streaming in both cases."""
    if filename.lower().endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        yield from csv.reader(io.StringIO(text))
        return

    try:
        import openpyxl
    except ModuleNotFoundError:  # pragma: no cover - dependency is in requirements
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Spreadsheet support is not installed on this server",
        )
    try:
        book = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="That file could not be opened as a spreadsheet",
        )
    try:
        yield from book[book.sheetnames[0]].iter_rows(values_only=True)
    finally:
        book.close()


def parse(data: bytes, filename: str) -> _Parsed:
    """Read the file into memory as facts, rejecting rows with a reason."""
    out = _Parsed()
    rows = _iter_rows(data, filename)

    header = next(rows, None)
    if header is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="The file is empty"
        )
    columns = {}
    for index, cell in enumerate(header):
        key = _HEADERS.get(_norm_header(cell))
        if key and key not in columns:
            columns[key] = index
    missing = [c for c in ("region", "state", "pincode") if c not in columns]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"The sheet needs a {', '.join(missing)} column. "
                f"Found: {', '.join(str(c) for c in header if c)}"
            ),
        )

    def cell(row: tuple, key: str) -> str:
        index = columns.get(key)
        if index is None or index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        # Pincodes arrive as ints in a real spreadsheet, not text.
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    for number, row in enumerate(rows, start=2):
        if row is None:
            continue
        out.rows_read += 1
        if out.rows_read > MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"That file has more than {MAX_ROWS:,} rows",
            )

        region = cell(row, "region")
        state = cell(row, "state")
        district = cell(row, "district")
        code = cell(row, "pincode")

        # A row the SOURCE could not resolve. Dropped, not rejected: it is not a
        # data error the uploader can fix row by row, and 715 of them would bury
        # the rejects that do matter.
        if region.lower() in _SOURCE_ERRORS or state.lower() in _SOURCE_ERRORS:
            out.rows_skipped += 1
            if PINCODE_RE.match(code):
                out.unresolved_pins.add(code)
            continue

        if not code:
            out.rejects.append(ImportReject(row=number, reason="No pincode"))
            continue
        if not PINCODE_RE.match(code):
            out.rejects.append(
                ImportReject(
                    row=number,
                    pincode=code,
                    reason="Not a 6-digit pincode (and none start with 0)",
                )
            )
            continue

        out.state_region.setdefault(state, region)
        out.pin_states[code][state] += 1
        if district and district.lower() not in _SOURCE_ERRORS:
            out.pin_districts[code].add(district)
            out.state_districts[state].add(district)

    return out


def resolve_states(parsed: _Parsed) -> tuple[dict[str, str], list[ImportOverride]]:
    """Decide the one state each pincode belongs to. Majority of rows wins.

    An exact tie is REJECTED by name rather than guessed at. There is nothing
    else: the spreadsheet is the only source, and a hard-coded correction that
    silently outranked it would mean fixing the file no longer fixed the master.
    Anything wrong is fixed in the sheet and re-uploaded.

    `overrides` is always empty now. It stays in the signature because the
    report carries the field and both clients read it, and because a future
    correction mechanism would report through it.
    """
    chosen: dict[str, str] = {}
    overrides: list[ImportOverride] = []

    for code, votes in parsed.pin_states.items():
        ranked = votes.most_common()
        if len(ranked) > 1 and ranked[0][1] == ranked[1][1]:
            parsed.rejects.append(
                ImportReject(
                    pincode=code,
                    reason=(
                        "Listed under "
                        + " and ".join(
                            f"{title_case(s)} ({n} row{'' if n == 1 else 's'})"
                            for s, n in ranked
                        )
                        + " with no majority — decide which and re-upload"
                    ),
                )
            )
            continue
        chosen[code] = ranked[0][0]

    # Pincodes the source could not resolve AND that appear nowhere else. Named
    # one by one: "715 rows skipped" does not tell anybody which addresses just
    # became unservable.
    for code in sorted(parsed.unresolved_pins - set(chosen)):
        parsed.rejects.append(
            ImportReject(
                pincode=code,
                reason=(
                    "Source lookup failed (#N/A) and this pincode has no state "
                    "anywhere else in the file"
                ),
            )
        )

    return chosen, overrides


# ── writing ────────────────────────────────────────────────────────────────


@dataclass
class _Existing:
    """What is already in the database, keyed the way the file will be matched."""

    regions: dict[str, Region] = field(default_factory=dict)          # lower(name)
    states: dict[str, State] = field(default_factory=dict)            # lower(name)
    districts: dict[tuple[uuid.UUID, str], District] = field(default_factory=dict)
    pincodes: dict[str, uuid.UUID] = field(default_factory=dict)      # code -> state_id
    links: set[tuple[str, uuid.UUID]] = field(default_factory=set)


async def _load_existing(session: AsyncSession) -> _Existing:
    out = _Existing()
    for region in (await session.scalars(select(Region))).all():
        out.regions[region.name.lower()] = region
    for state in (await session.scalars(select(State))).all():
        out.states[state.name.lower()] = state
    for district in (await session.scalars(select(District))).all():
        out.districts[(district.state_id, district.name.lower())] = district
    for code, state_id in await session.execute(
        select(Pincode.code, Pincode.state_id)
    ):
        out.pincodes[code] = state_id
    for code, district_id in await session.execute(
        select(PincodeDistrict.pincode_code, PincodeDistrict.district_id)
    ):
        out.links.add((code, district_id))
    return out


async def import_geography(
    session: AsyncSession,
    data: bytes,
    filename: str,
    *,
    dry_run: bool,
    actor_id: uuid.UUID | None,
) -> ImportReport:
    """Load the file. Additive: creates and updates, never deletes what is absent."""
    parsed = parse(data, filename)
    chosen, overrides = resolve_states(parsed)
    existing = await _load_existing(session)

    regions = ImportCounts()
    states = ImportCounts()
    districts = ImportCounts()
    pincodes = ImportCounts()

    # Region names the file uses, resolved to rows. `sort_order` continues the
    # existing sequence so a new region lands after the seeded five.
    next_sort = max((r.sort_order for r in existing.regions.values()), default=0) + 10
    region_by_name: dict[str, Region] = {}
    for region_name in dict.fromkeys(parsed.state_region.values()):
        key = region_name.lower()
        row = existing.regions.get(key)
        if row is not None:
            regions.updated += 1
        else:
            row = Region(
                id=uuid.uuid4(),
                code=re.sub(r"[^A-Z0-9]", "", region_name.upper())[:16] or "REGION",
                name=title_case(region_name),
                sort_order=next_sort,
                is_active=True,
                created_by=actor_id,
            )
            next_sort += 10
            existing.regions[key] = row
            regions.created += 1
            if not dry_run:
                session.add(row)
        region_by_name[key] = row

    if not dry_run and regions.created:
        await session.flush()  # autoflush is off; ids are needed below

    # States.
    state_rows: dict[str, State] = {}
    for state_name, region_name in parsed.state_region.items():
        region = region_by_name[region_name.lower()]
        key = state_name.lower()
        row = existing.states.get(key)
        if row is None:
            row = State(
                id=uuid.uuid4(),
                region_id=region.id,
                name=title_case(state_name),
                is_active=True,
                created_by=actor_id,
            )
            existing.states[key] = row
            states.created += 1
            if not dry_run:
                session.add(row)
        elif row.region_id != region.id:
            # A state changing region moves every pincode under it, and with
            # them whichever regional head sees the work. Never silent.
            states.moved += 1
            if not dry_run:
                row.region_id = region.id
                row.updated_by = actor_id
        else:
            states.updated += 1
        state_rows[key] = row

    if not dry_run and states.created:
        await session.flush()

    # Districts, keyed by (state, name) because five names repeat across states.
    district_rows: dict[tuple[str, str], District] = {}
    for state_name, names in parsed.state_districts.items():
        state = state_rows[state_name.lower()]
        for name in names:
            key = (state.id, name.lower())
            row = existing.districts.get(key)
            if row is None:
                row = District(
                    id=uuid.uuid4(),
                    state_id=state.id,
                    name=title_case(name),
                    created_by=actor_id,
                )
                existing.districts[key] = row
                districts.created += 1
                if not dry_run:
                    session.add(row)
            else:
                districts.updated += 1
            district_rows[(state_name.lower(), name.lower())] = row

    if not dry_run and districts.created:
        await session.flush()

    # Pincodes.
    moved: list[dict] = []
    for code, state_name in chosen.items():
        state = state_rows[state_name.lower()]
        current = existing.pincodes.get(code)
        if current is None:
            pincodes.created += 1
            if not dry_run:
                session.add(
                    Pincode(
                        code=code,
                        state_id=state.id,
                        is_active=True,
                        created_by=actor_id,
                    )
                )
        elif current != state.id:
            pincodes.moved += 1
            if not dry_run:
                # Collected, not fetched one at a time: a re-import that moves
                # thousands of codes would otherwise be thousands of round
                # trips and time the request out.
                moved.append({"c": code, "s": state.id})
        else:
            pincodes.updated += 1

    if not dry_run and moved:
        await session.execute(
            update(Pincode)
            .where(Pincode.code == bindparam("c"))
            .values(state_id=bindparam("s"), updated_by=actor_id),
            moved,
        )
    if not dry_run and (pincodes.created or pincodes.moved):
        await session.flush()

    # Pincode -> district links. The file is authoritative for the pincodes it
    # names, so their links are replaced; pincodes it does not name are left
    # alone entirely.
    if not dry_run:
        wanted: set[tuple[str, uuid.UUID]] = set()
        for code, state_name in chosen.items():
            for name in parsed.pin_districts.get(code, ()):
                district = district_rows.get((state_name.lower(), name.lower()))
                if district is not None:
                    wanted.add((code, district.id))

        touched = set(chosen)
        stale = {
            link for link in existing.links if link[0] in touched and link not in wanted
        }
        for code, district_id in stale:
            link = await session.get(PincodeDistrict, (code, district_id))
            if link is not None:
                await session.delete(link)
        for code, district_id in sorted(wanted - existing.links):
            session.add(
                PincodeDistrict(
                    pincode_code=code, district_id=district_id, created_by=actor_id
                )
            )
        await session.commit()
    else:
        # Nothing was written, but rows were added to the identity map above
        # only when `dry_run` is false. Expire anyway so a later read in the
        # same session cannot see a half-built object.
        session.expunge_all()

    used = {r.lower() for r in parsed.state_region.values()}
    unused = sorted(
        r.name for key, r in existing.regions.items() if key not in used and r.is_active
    )

    return ImportReport(
        dryRun=dry_run,
        rowsRead=parsed.rows_read,
        rowsSkipped=parsed.rows_skipped,
        regions=regions,
        states=states,
        districts=districts,
        pincodes=pincodes,
        unusedRegions=unused,
        overrides=overrides,
        rejected=len(parsed.rejects),
        rejects=parsed.rejects[:MAX_REJECTS_RETURNED],
    )
