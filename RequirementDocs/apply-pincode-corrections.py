"""Write the researched corrections INTO `Reliance Green Tech Pin Code.xlsx`.

    python RequirementDocs/apply-pincode-corrections.py [--write]

The sheet is the single source of truth. There are no overrides in the importer
any more, so anything wrong in the file is wrong in the master — which means a
correction has to be made here, in the data, not hidden in code.

Every change below was researched against India Post; `Pin Code corrections.md`
records the evidence for each one. Run with `--write` to apply; without it the
script only reports.

The original is in git (`git show a2d8f65:"RequirementDocs/Reliance Green Tech
Pin Code.xlsx"`), so this is reversible.
"""

import sys
from pathlib import Path

import openpyxl

SRC = Path(__file__).with_name("Reliance Green Tech Pin Code.xlsx")

#: pincode -> {row exactly as it reads now: what it should read}.
#: A row whose triple is not listed is left untouched.
REWRITE: dict[str, dict[tuple, tuple]] = {
    # -- the source lookup failed; region/state/district all blank ------------
    "222101": {("#N/A", "NA", "NA"): ("North", "UTTAR PRADESH", "JAUNPUR")},
    "390008": {("#N/A", "NA", "NA"): ("West", "GUJARAT", "VADODARA")},
    "605012": {("#N/A", "NA", "NA"): ("South", "PUDUCHERRY", "PONDICHERRY")},
    "804454": {("#N/A", "NA", "NA"): ("East", "BIHAR", "PATNA")},
    # -- wrong district: every India Post office here is in Bijapur -----------
    "494446": {
        ("West", "CHHATTISGARH", "RAIPUR"): ("West", "CHHATTISGARH", "BIJAPUR"),
        ("#N/A", "NA", "NA"): ("West", "CHHATTISGARH", "BIJAPUR"),
    },
    # -- a pincode straddling a border, so the file could not decide ----------
    # The master stores ONE state per pincode, so the rows are made to agree
    # with the administrative answer instead of splitting the vote.
    "605014": {
        ("South", "TAMIL NADU", "VILLUPURAM"): ("South", "PUDUCHERRY", "PONDICHERRY"),
    },
    "605106": {
        ("South", "TAMIL NADU", "VILLUPURAM"): ("South", "PUDUCHERRY", "PONDICHERRY"),
        ("South", "TAMIL NADU", "CUDDALORE"): ("South", "PUDUCHERRY", "PONDICHERRY"),
    },
    "605107": {
        ("South", "PUDUCHERRY", "PONDICHERRY"): ("South", "TAMIL NADU", "VILLUPURAM"),
    },
    "607403": {
        ("South", "PUDUCHERRY", "PONDICHERRY"): ("South", "TAMIL NADU", "CUDDALORE"),
    },
    "781029": {
        ("East", "MEGHALAYA", "RI BHOI"): ("East", "ASSAM", "KAMRUP METRO"),
    },
    "781131": {
        ("East", "MEGHALAYA", "RI BHOI"): ("East", "ASSAM", "KAMRUP"),
    },
}

#: Live delivery pincodes India Post has and this export never mentioned.
#: 393155 gets two rows because it genuinely spans two districts.
ADD: list[tuple[str, str, str, int]] = [
    ("West", "RAJASTHAN", "GANGANAGAR", 335705),
    ("West", "GUJARAT", "RAJKOT", 364485),
    ("West", "GUJARAT", "NARMADA", 393155),
    ("West", "GUJARAT", "BHARUCH", 393155),
    ("West", "GUJARAT", "NAVSARI", 396424),
    ("West", "GUJARAT", "NAVSARI", 396440),
    ("East", "BIHAR", "PASHCHIM CHAMPARAN", 845102),
]


def cell(v):
    return "" if v is None else str(v).strip()


def main(write: bool) -> int:
    book = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    source = book[book.sheetnames[0]]
    rows = source.iter_rows(values_only=True)
    header = next(rows)

    # Only built when actually writing. A write-only workbook streams to a
    # temp file, and one left unsaved is reaped at exit — which on Windows
    # throws a PermissionError over the top of the report.
    out = sheet = None
    if write:
        out = openpyxl.Workbook(write_only=True)
        sheet = out.create_sheet("Pincodes")
        sheet.append(list(header))

    changed = 0
    per_code: dict[str, int] = {}
    total = 0

    for row in rows:
        total += 1
        region, state, district = (cell(row[i]) for i in range(3))
        raw = row[3] if len(row) > 3 else None
        code = ""
        if raw is not None:
            code = str(int(raw)) if isinstance(raw, float) and raw.is_integer() else cell(raw)

        swap = REWRITE.get(code, {}).get((region, state, district))
        if swap:
            region, state, district = swap
            changed += 1
            per_code[code] = per_code.get(code, 0) + 1

        if sheet is not None:
            sheet.append([region, state, district, raw])

    if sheet is not None:
        for region, state, district, code in ADD:
            sheet.append([region, state, district, code])

    book.close()

    print(f"rows read     : {total:,}")
    print(f"rows rewritten: {changed}")
    for code in sorted(per_code):
        print(f"   {code}  x{per_code[code]}")
    print(f"rows added    : {len(ADD)}")
    for region, state, district, code in ADD:
        print(f"   {code}  {region} / {state} / {district}")
    print(f"rows out      : {total + len(ADD):,}")

    if not write:
        print("\nNothing written. Re-run with --write.")
        return 0

    out.save(SRC)
    print(f"\nwrote {SRC}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main("--write" in sys.argv))
